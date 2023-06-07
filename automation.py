import openpyxl as pyxl
import shutil as stil

filename = "Salary_Data.xlsx"
backup_filename = "Salary_Data_backup.xlsx"
stil.copyfile(filename, backup_filename)

def automation(filename):
     wb = pyxl.load_workbook(filename)
     sheet = wb['Salary_Data']


     for row in range(2, sheet.max_row+1):
         sheet["c1"] = "Updated salary"
         cell = sheet.cell(row, 2)
         correct_salary = cell.value * 2
         correct_cell = sheet.cell(row, 3)
         correct_cell.value = correct_salary

     wb.save(filename)

automation(filename)


